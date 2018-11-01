# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import sqlite3


def set_border():
    bd = Side(style='thin', color="000000")  #设置单元格边框为细线，颜色为黑色
    border = Border(left=bd, top=bd, right=bd, bottom=bd)  #设置上下左右边框
    return border


kh_book = Workbook()
kh_sheet = kh_book.active
#标题
kh_sheet.title = '外县'
kh_sheet.merge_cells('A1:F1')
kh_sheet['A1'] = '计算机协管员考核表'
kh_sheet['A1'].font = Font(name='微软雅黑', size=14, bold=True)
kh_sheet['A1'].border = set_border()
kh_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
kh_sheet.row_dimensions[1].height = 30
#表头
kh_lables = [r'序号', r'员工编号代码', r'姓名代码', r'所在机构代码', r'所在部门代码', r'激励费']
kh_col = ['A', 'B', 'C', 'D', 'E', 'F']
kh_sheet.column_dimensions['A'].width = 8
kh_sheet.column_dimensions['B'].width = 20
kh_sheet.column_dimensions['C'].width = 10
kh_sheet.column_dimensions['D'].width = 16
kh_sheet.column_dimensions['E'].width = 16
kh_sheet.column_dimensions['F'].width = 14
row_num = 2
for lable, col in zip(kh_lables, kh_col):
    kh_sheet[col + str(row_num)] = lable
    kh_sheet[col + str(row_num)].border = set_border()
    kh_sheet[col + str(row_num)].alignment = Alignment(horizontal='center')
    kh_sheet[col + str(row_num)].font = Font(name='微软雅黑', size=11)
#数据
conn = sqlite3.connect('itsm.db')
cursor = conn.cursor()
sql = 'select * from report;'
data = cursor.execute(sql)
values = data.fetchall()
num = 1
total = 0
for value in values:
    row_num += 1
    kh_sheet['A' + str(row_num)] = num
    kh_sheet['B' + str(row_num)] = value[1]
    kh_sheet['C' + str(row_num)] = value[0]
    kh_sheet['D' + str(row_num)] = value[4]
    kh_sheet['E' + str(row_num)] = value[2]
    jine = 0
    if value[4] == '延吉':
        if value[5] >= 10:
            jine = 200
        else:
            jine = 200 - (10 - value[5]) * 10
    else:
        if value[5] >= 15:
            jine = 300
        else:
            jine = 300 - (15 - value[5]) * 10
        jine = jine - value[6] * 20
    kh_sheet['F' + str(row_num)] = jine
    total = total + jine
    for col in kh_col:
        kh_sheet[col + str(row_num)].border = set_border()
        kh_sheet[col + str(row_num)].alignment = Alignment(horizontal='center')
        kh_sheet[col + str(row_num)].font = Font(name='微软雅黑', size=11)
    num += 1
#最后合计的一行
row_num += 1
kh_sheet['A' + str(row_num)] = r'合计'
kh_sheet['B' + str(row_num)] = total
for col in kh_col:
    kh_sheet[col + str(row_num)].border = set_border()
    kh_sheet[col + str(row_num)].alignment = Alignment(horizontal='center')
    kh_sheet[col + str(row_num)].font = Font(name='微软雅黑', size=11)

kh_sheet.merge_cells('A%s:F%s' % (str(row_num + 1), str(row_num + 1)))
kh_sheet['A' + str(row_num +
                   1)] = '人力资源部           财务部负责人          本部门负责人          制表人'
kh_book.save('report.xlsx')
